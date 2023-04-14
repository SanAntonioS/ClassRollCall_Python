import os
import openpyxl #一个处理excel的库
import random   #生成随机数据的库

#Qt库-实现UI界面
from PySide6.QtWidgets import QApplication, QMainWindow, QPushButton,  QPlainTextEdit, QMessageBox, QFileDialog, QTableWidgetItem
from PySide6.QtUiTools import QUiLoader
from PySide6.QtCore import QDir, Qt

import win32com.client as win32 #格式转换库
from datetime import datetime   #获取时间库
import pyttsx3                  #语言播报库

studentNum = 0

#将xls格式文件转换为xlsx格式的"Log.xlsx"，openpyxl包无法处理xls格式文件
def xls2xlsx(filePath):
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    wb = excel.Workbooks.Open(filePath)
    wb.SaveAs(filePath[:-38]+"Log.xlsx", FileFormat = 51)   #FileFormat = 51 is for .xlsx extension
    # wb.SaveAs(fname[:-1], FileFormat = 56)                #FileFormat = 56 is for .xls extension
    wb.Close()
    excel.Application.Quit()

#处理转换后的"Log.xlsx"
def ReadData():
    workbook = openpyxl.load_workbook("Log.xlsx")	#返回一个workbook数据类型的值
    sheet = workbook.active                         
    student_id = sheet['A3:F49']                    #将表格中的数据读取到组中
    student_id_list = list(student_id)              #将组转换成列表
    random.shuffle(student_id_list)                 #将列表打乱顺序，按行打乱
    workbook.close()
    
    return student_id_list                          #返回一个乱序的列表
    
class Stats():

    #初始化Qt的UI界面
    def __init__(self):
        self.ui = QUiLoader().load('main.ui')
        self.ui.tableWidget.setColumnWidth(0, 70)
        self.ui.tableWidget.setColumnWidth(1, 120)
        self.ui.tableWidget.setColumnWidth(2, 120)
        self.ui.RollCall.clicked.connect(self.RollCall)
        self.ui.Absenteeism.clicked.connect(self.Absenteeism)
        self.ui.ImportList.clicked.connect(self.ImportList)
        self.ui.Sort.clicked.connect(self.Sort)
        self.ui.SaveData.clicked.connect(self.SaveData)
        self.ui.OpenData.clicked.connect(self.OpenData)

    #槽函数：点名
    def RollCall(self):
        global studentNum
        
        #显示在左边三栏上
        self.ui.studentName.setPlainText(student_id_list[studentNum][2].value)
        self.ui.studentClass.setPlainText(student_id_list[studentNum][0].value)
        self.ui.studentID.setPlainText(student_id_list[studentNum][1].value)
        
        #播放被点到的人的姓名
        speaker.say(student_id_list[studentNum][2].value)
        speaker.runAndWait()
        
        #更新右边表格数据
        self.ui.tableWidget.insertRow(0)
        nameItem = QTableWidgetItem("%s" % student_id_list[studentNum][2].value)
        classItem = QTableWidgetItem("%s" % student_id_list[studentNum][0].value)
        idItem = QTableWidgetItem("%s" % student_id_list[studentNum][1].value)
        statusItem = QTableWidgetItem("%s" % '已到')
        self.ui.tableWidget.setItem(0,0,nameItem)
        self.ui.tableWidget.setItem(0,1,classItem)
        self.ui.tableWidget.setItem(0,2,idItem)
        self.ui.tableWidget.setItem(0,3,statusItem)
        
        #计入标志位实现乱序且不重复的点名
        studentNum += 1
        
    def Sort(self):
        #按学号排序
        self.ui.tableWidget.sortItems(2,Qt.AscendingOrder)
    
    def Absenteeism(self):
        #将被点到的学生状态修改成：旷课
        statusItem = QTableWidgetItem("%s" % '旷课')
        self.ui.tableWidget.setItem(0,3,statusItem)
        
    def ImportList(self):
        #导入名单，返回该名单的路径
        filePath = QFileDialog.getOpenFileName(self.ui, "导入名单")[0]
        filePath = QDir.toNativeSeparators(filePath)
        print(filePath)
        xls2xlsx(filePath)
    
    def SaveData(self):
        #将点名的数据保存至"Log.xlsx"
        workbook = openpyxl.load_workbook("Log.xlsx")
        sheet = workbook.active
        sheet.insert_cols(7)
        sheet['G2'] = datetime.now().strftime('%Y-%m-%d')   #插入本次保存的时间
        
        tableRow = self.ui.tableWidget.rowCount()
        for i in range(tableRow):
            tableName = self.ui.tableWidget.item(i,0).text()
            for row in range(3,49):
                name = sheet.cell(row ,3).value
                if name == tableName:
                    sheet.cell(row, 7, self.ui.tableWidget.item(i,3).text())
        workbook.save("Log.xlsx")
        workbook.close()
        
    def OpenData(self):
        #打开保存的点名册
        os.startfile("Log.xlsx")
    

speaker = pyttsx3.init()    #初始化语言播报模块
msg = "你好"
speaker.say(msg)
speaker.runAndWait()

app = QApplication([])      #初始化Qt槽函数
stats = Stats()
stats.ui.show()

student_id_list = ReadData()#读取被打乱的学生名单

app.exec()                  #等待槽函数的执行
